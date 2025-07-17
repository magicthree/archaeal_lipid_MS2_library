import os
from openpyxl import load_workbook
from MSPfix import fix_file

def import_adduct():
    return [["[M+H]+", 1.007276], ["[M+NH4]+", 18.03382]]

def single_msp_export(folder_path,file_name,export_path=None,newdir=None,fix=True,max_peak=1000,min_peak=0.1):
    try:
        adducts = import_adduct()
        if export_path is None:
            export_path=folder_path
        if newdir is None:
            newdir= file_name + '.msp'

        file_path = os.path.join(folder_path, file_name)
        workbook = load_workbook(file_path, data_only=True)
        print(newdir+" file loaded")
        f = open(export_path + '\\' + newdir, 'w')


        sheet_names = workbook.sheetnames
        sheet = workbook['ExportControl']
        min_value=999
        max_value=1
        for cell in sheet['A']:
            cell_value = cell.value
            if cell_value is not None and int(cell_value) > 0:
                min_value = min(min_value, cell_value)
                max_value = max(max_value, cell_value)

        for i in range(min_value-1, max_value):

            sheet = workbook[sheet_names[i]]
            molecule_count = 0

            for row in range(7, 2000 + 1):
                cell_value = sheet["A"+str(row)].value
                if cell_value is not None:
                    molecule_count += 1
            peak_count=sheet["B1"].value
            for adduct in adducts:
                name_list=[]
                for j in range (1,molecule_count+1):
                    name_list.append(sheet["N" + str(j + 6)].value)
                for j in range(1,molecule_count+1):
                    confidence_level=sheet["M" + str(j + 6)].value
                    Name=name_list[j-1]
                    f.write("Name: " + str(Name) + '\n' +
                            "MW: " + str(round(float(sheet["B" + str(j + 6)].value), 5)) + '\n' +
                            "PRECURSORMZ: " + str(round(float(sheet["B" + str(j + 6)].value) + adduct[1], 5)) + '\n' +
                            "Formula: " + str(sheet["D" + str(j + 6)].value) + '\n' +
                            "RETENTIONTIME: " + str(sheet["I" + str(j + 6)].value) + '\n' +
                            "logP: " + str(round(float(sheet["J" + str(j + 6)].value), 5)) + '\n' +
                            "INCHI: " + str(sheet["F" + str(j + 6)].value) + '\n' +
                            "INCHIKEY: " + str(sheet["K" + str(j + 6)].value) + '\n' +
                            "SMILES: " + str(sheet["G" + str(j + 6)].value) + '\n' +
                            "CCS: " + str(sheet["H" + str(j + 6)].value) + '\n' +
                            "PRECURSORTYPE: " + str(adduct[0]) + '\n' +
                            "Comment: " + str(sheet["C2"].value) + '\n' +
                            "Confidencelevel: " + str(confidence_level) + '\n' +
                            "Num Peaks: " + str(peak_count) + '\n')
                    for peak in range(0,peak_count):
                        f.write(str(sheet.cell(row=j+6,column=29+peak).value)+" "+str(sheet.cell(row=j+6,column=29+peak+peak_count).value)+" "+chr(34)+str(sheet.cell(row=6,column=29+peak).value)+chr(34)+'\n')
                    f.write('\n')
        workbook.close()
        print(newdir+" file.msp made")

        f.close()
        if fix:
            fix_file(export_path,newdir,max_peak,min_peak)
            print(newdir + " file.msp fixed")
    except Exception as e:
        print(file_name + "###################failed#################" + e)
    return newdir+'.msp'

